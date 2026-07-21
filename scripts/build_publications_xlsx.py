#!/usr/bin/env python3
"""Build publications.xlsx from data/publications_clean.json.

Usage:
    python3 build_publications_xlsx.py [data_dir]

Writes data/publications.xlsx with columns:
Title, Year, Venue, Type, Co-authors
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADERS = ["Title", "Year", "Venue", "Type", "Co-authors"]


def main():
    data_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent / "data"
    pubs = json.loads((data_dir / "publications_clean.json").read_text())

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT, bold=True)

    for row, pub in enumerate(pubs, start=2):
        ws.cell(row=row, column=1, value=pub["title"]).font = Font(name=FONT)
        ws.cell(row=row, column=2, value=pub["year"]).font = Font(name=FONT)
        ws.cell(row=row, column=3, value=pub["venue"]).font = Font(name=FONT)
        ws.cell(row=row, column=4, value=pub["type"]).font = Font(name=FONT)
        ws.cell(row=row, column=5, value="; ".join(pub["co_authors"])).font = Font(name=FONT)

    widths = {1: 70, 2: 8, 3: 35, 4: 14, 5: 60}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    out_path = data_dir / "publications.xlsx"
    wb.save(out_path)
    print(f"Wrote {len(pubs)} publications -> {out_path}")


if __name__ == "__main__":
    main()
